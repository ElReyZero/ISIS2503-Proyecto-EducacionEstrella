import sys
import psycopg2 as pg
import csv
import traceback
from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Font
import os
from datetime import datetime
from multiprocessing import Process, Manager


def connect():
    manager = Manager()
    queries = manager.dict()
    procesoCarrera = Process(target=queryCarrera, args=(queries,))
    procesoUniversidad = Process(target=queryUniversidad, args=(queries,))
    procesoCiudad = Process(target=queryCiudad, args=(queries,))
    procesoGenero = Process(target=queryGenero, args=(queries,))
    procesoEdad = Process(target=queryEdad, args=(queries,))

    procesoCarrera.start()
    procesoUniversidad.start()
    procesoCiudad.start()
    procesoGenero.start()
    procesoEdad.start()

    procesoCarrera.join()
    procesoUniversidad.join()
    procesoCiudad.join()
    procesoGenero.join()
    procesoEdad.join()

    queryToExcel(queries)


def queryTasaDeAprobacion(criterio):

    """El criterio debe ser:
        -carrera
        -ciudad
        -universidad
        -genero
        -edad
    """
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        query = "SELECT b."+criterio+", "
        query+= "CASE WHEN a.creditos_aprobados IS NOT NULL THEN ROUND((a.creditos_aprobados/b.creditos_totales),2) "
        query += "ELSE 0 END AS tasa_aprobacion "
        query += " FROM ( "
        query += " SELECT CAST (COUNT(*) as numeric(9,2)) creditos_aprobados, ee."+criterio+" "
        query += " FROM public.\"ModuloFinanciero_solicitudcredito\" s, public \"ModuloFinanciero_estudianteestrella\" ee "
        query += " WHERE s.estudiante_id = ee.id AND s.\"fechaAprobacion\" IS NOT NULL "
        query += " GROUP BY ee."+criterio+") "
        query += " a RIGHT JOIN ( "
        query += " SELECT CAST (COUNT(*) as numeric(9,2)) creditos_totales, ee."+criterio +" "
        query += " FROM public.\"ModuloFinanciero_solicitudcredito\" s,public.\"ModuloFinanciero_estudianteestrella\" ee "
        query += " WHERE s.estudiante_id = ee.id "
        query += " GROUP BY ee."+criterio +" ) b ON a."+criterio+" = b."+criterio+";"

        cursor.execute(query)
        data = cursor.fetchall()
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryRecaudoSemestral(criterio):

    """El criterio debe ser:
        -carrera
        -ciudad
        -universidad
        -genero
        -edad
    """
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        query = "SELECT "+criterio+", SUM(\"montoAPagar\") "
        query += " FROM public.\"ModuloFinanciero_solicitucredito\" s, public.\"ModuloFinanciero_estudianteestrella\" ee "
        query += " WHERE \"fechaAprobacion\" BETWEEN TO_DATE('2002-01-01', 'YYYY-MM-DD' AND "
        query += " TO_DATE('2022-06-06', 'YYYY-MM-DD') AND pagado AND s.estudiante_id = ee.id "
        query += " GROUP BY ee."+criterio+";"

        cursor.execute(query)
        data = cursor.fetchall()
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryCuotasEnMora(criterio):

    """El criterio debe ser:
        -carrera
        -ciudad
        -universidad
        -genero
        -edad
    """
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        query = "SELECT "+criterio +", COUNT(*) "
        query += " FROM public.\"ModuloFinanciero_solicitudcredito\" s, public.\"ModuloFinanciero_estudianteestrella\" ee"
        query += " WHERE NOT pagado AND s.estudiante_id = ee.id "
        query += " GROUP BY ee."+criterio+";"
        
        cursor.execute(query)
        data = cursor.fetchall()
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryCarrera(queries):
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        cursor.execute("SELECT * FROM public.\"ModuloFinanciero_estudianteestrella\"")
        data = cursor.fetchall()
        queries["Carreras"] = data
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

            

def queryUniversidad(queries):
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        cursor.execute("SELECT * FROM public.\"ModuloFinanciero_estudianteestrella\"")
        data = cursor.fetchall()
        queries["Universidades"] = data
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryCiudad(queries):
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        cursor.execute("SELECT * FROM public.\"ModuloFinanciero_estudianteestrella\"")
        data = cursor.fetchall()
        queries["Ciudades"] = data
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryGenero(queries):
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        cursor.execute("SELECT * FROM public.\"ModuloFinanciero_estudianteestrella\"")
        data = cursor.fetchall()
        queries["Género"] = data
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryEdad(queries):
    try:
        conn = pg.connect(
            host = "educacion-estrella-db.cgg09hgjbyvt.us-east-1.rds.amazonaws.com",
            database = "educacionEstrellaDB",
            user = "eeUser",
            password = "isis2503"
            )
        cursor = conn.cursor()
        print("Connected to database")

        cursor.execute("SELECT * FROM public.\"ModuloFinanciero_estudianteestrella\"")
        data = cursor.fetchall()
        queries["Edad"] = data
        cursor.close()
        conn.close()
    except (Exception, pg.DatabaseError):
        print(traceback.format_exc())
    except (KeyboardInterrupt):
        if conn:
            conn.close()
        sys.exit()
    finally:
        if conn:
            conn.close()

def queryToExcel(queries:dict):
    headings = ("col1", "col2", "col3", "col4", "col5", "col6", "col7")
    wb = Workbook()
    sheet = wb.active
    last = list(queries.keys())[-1]
    for name, query in queries.items():
        sheet.title = name
        sheet.row_dimensions[1].font = Font(bold = True)
        # Spreadsheet row and column indexes start at 1
        # so we use "start = 1" in enumerate so
        # we don't need to add 1 to the indexes.
        for colno, heading in enumerate(headings, start = 1):
            sheet.cell(row = 1, column = colno).value = heading

        # This time we use "start = 2" to skip the heading row.
        for rowno, row in enumerate(query, start = 2):
            for colno, cell_value in enumerate(row, start = 1):
                try:
                    if isinstance(cell_value, datetime):
                        cell_value = datetime.strftime(cell_value, "%d/%m/%Y")
                except TypeError:
                    pass
                sheet.cell(row = rowno, column = colno).value = cell_value
        if not name == last:
            sheet = wb.create_sheet(f"Hoja {name}")
    try:
        os.makedirs("./reports")
    except FileExistsError:
        pass
    wb.security = WorkbookProtection(workbookPassword="pass", revisionsPassword="pass", lockStructure=True, lockRevision=True)
    wb.save(f"./reports/reporte.xlsx")


def insertEstudiantes(conn, cur):
    with open(r'C:\Users\ElRey\Documents\Scripts\ISIS2503-Proyecto-EducacionEstrella\ReportGatherer\estudiantes.csv', 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        cont = 0
        for row in reader:
            if len(row[4]) > 50:
                row[4] = row[4][:50]
            cur.execute(
            'INSERT INTO public."ModuloFinanciero_estudianteestrella" VALUES (%s, %s, %s, %s, %s, %s, %s)',
            row)
            cont += 1
            print(f"\r{cont}/1000", end="")
        conn.commit()
        cur.close()
        print("\nCarga finalizada")

def insertSolicitudes(conn, cur):
    with open(r'C:\Users\ElRey\Documents\Scripts\ISIS2503-Proyecto-EducacionEstrella\ReportGatherer\solicitudes.csv', 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        cont = 0
        for row in reader:
            if row[3] == "":
                row[3] = None
                row[6] = 'false'
            cur.execute(
            'INSERT INTO public."ModuloFinanciero_solicitudcredito" VALUES (%s, %s, %s, %s, %s, %s, %s)', row)
            cont += 1
            print(f"\r{cont}/1000", end="")
        conn.commit()
        cur.close()
        print("\nCarga finalizada")

if __name__ == '__main__':
    connect()